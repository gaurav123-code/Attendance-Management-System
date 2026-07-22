import csv

import openpyxl

from django.http import HttpResponse
from django.shortcuts import render

from ..decorators import admin_required
from ..models import Attendance
from django.utils import timezone
from openpyxl.utils import get_column_letter


# ==========================
# Export Page
# ==========================

@admin_required
def export_page(request):

    return render(
        request,
        "attendance/export_report.html"
    )


# ==========================
# Export Attendance CSV
# ==========================

@admin_required
def export_attendance_csv(request):

    response = HttpResponse(
        content_type="text/csv"
    )

    response["Content-Disposition"] = (
        'attachment; filename="attendance_report.csv"'
    )

    writer = csv.writer(response)

    writer.writerow([
        "Employee ID",
        "Employee Name",
        "Department",
        "Date",
        "Status",
        "Check In",
        "Check Out",
        "Working Hours"
    ])

    selected_date = request.GET.get("date")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")


    attendance_records = Attendance.objects.select_related(
        "employee",
        "employee__department"
    ).filter(
        employee__is_active=True
    )


    if selected_date:

        attendance_records = attendance_records.filter(
            attendance_date=selected_date
        )


    elif start_date and end_date:

        attendance_records = attendance_records.filter(
            attendance_date__range=[
                start_date,
                end_date
            ]
        )


    for attendance in attendance_records:

        writer.writerow([

            attendance.employee.employee_id,

            (
                attendance.employee.first_name
                +
                " "
                +
                attendance.employee.last_name
            ),

            attendance.employee.department.name,

            attendance.attendance_date.strftime("%d-%m-%Y"),

            attendance.status,

            attendance.check_in,

            attendance.check_out,

            attendance.working_hours,

        ])


    return response





# ==========================
# Export Attendance Excel
# ==========================

@admin_required
def export_attendance_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Attendance Report"


    sheet.append([

        "Employee ID",
        "Employee Name",
        "Department",
        "Date",
        "Status",
        "Check In",
        "Check Out",
        "Working Hours"

    ])



    selected_date = request.GET.get("date")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")


    attendance_records = Attendance.objects.select_related(
        "employee",
        "employee__department"
    ).filter(
        employee__is_active=True
    )


    if selected_date:

        attendance_records = attendance_records.filter(
            attendance_date=selected_date
        )


    elif start_date and end_date:

        attendance_records = attendance_records.filter(
            attendance_date__range=[
                start_date,
                end_date
            ]
        )


    for attendance in attendance_records:

        sheet.append([

            attendance.employee.employee_id,

            (
                attendance.employee.first_name
                +
                " "
                +
                attendance.employee.last_name
            ),

            attendance.employee.department.name,

            attendance.attendance_date.strftime("%d-%m-%Y"),

            attendance.status,

            str(attendance.check_in),

            str(attendance.check_out),

            str(attendance.working_hours),

        ])




    response = HttpResponse(

        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )
    
    # Auto adjust column width

    for column_cells in sheet.columns:

        length = max(
            len(str(cell.value))
            if cell.value
            else 0
            for cell in column_cells
        )

        sheet.column_dimensions[
            get_column_letter(
                column_cells[0].column
            )
        ].width = length + 3


    response["Content-Disposition"] = (

        'attachment; filename="attendance_report.xlsx"'

    )



    workbook.save(response)


    return response





# ==========================
# Export Report Wrapper
# ==========================

@admin_required
def export_report(request):

    export_type = request.GET.get(
        "type"
    )


    if export_type == "csv":
        return export_attendance_csv(request)


    return export_attendance_excel(request)